# -*- coding: utf-8 -*-
from django.shortcuts import render
from django.shortcuts import HttpResponse, HttpResponseRedirect
from django.core.urlresolvers import reverse
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import redirect

from .models import SvtType, SvtUnit, SvtPeriod
from forms import PeriodForm
#from django import forms
import datetime


from openpyxl import load_workbook

# Create your views here.

def start(request):
    docs = SvtPeriod.objects.all()
    context = {'docs':docs}
    return render(request, 'documents/index.html', context)

def load_sap_file(request):
    if request.method == 'POST':
        form = PeriodForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            print "SAVED"
        else:
            print "NOT SAVED"
        context = {}
        #return render(request, 'documents/index.html', context)
        return redirect(reverse('start'))
    else:
        form = PeriodForm()

        context = {'form':form, }
        return render(request, 'documents/load_sap_file.html', context)


def load_to_db(request, doc_id):

    doc = SvtPeriod.objects.get(id=doc_id)
    #unit = SvtUnit()
    wb = load_workbook(doc.upload)
    sheet_ranges = wb[wb.sheetnames[0]]
    i=2

    unit = SvtUnit()
    while sheet_ranges['D'+str(i)].value is not None:

        print "Raw number: ", i

        unit.bar_code = sheet_ranges['D' + str(i)].value
        unit.type = sheet_ranges['A' + str(i)].value
        unit.sn = sheet_ranges['C' + str(i)].value
        unit.desc = sheet_ranges['B' + str(i)].value
        unit.save()
        i = i + 1

    context = {  }
    return render(request, 'documents/load_to_db.html', context)
